import openpyxl
import datetime
from openpyxl.styles import Font
from io import BytesIO

def creat_excel(df_select, house, name):
    font = Font(size=7)
    list1 = list(df_select["使用箇所"])
    list2 = list(df_select["使用部位"])
    list3 = list(df_select["メーカー"])
    list4 = list(df_select["商品名"])
    list5= list(df_select["型番"])
    list6 = list(df_select["色番号"])
    list7 = list(df_select["備考"])
    list_group = [list1, list2, list3, list4, list5, list6, list7]
    wb = openpyxl.load_workbook('./Data/仕上げリストテンプ.xlsx')
    ws = wb.active
    dt_now = datetime.datetime.now()
    list_row = ["A", "D", "H", "L", "S", "V", "AB"]
    for i in range(7):
        for j in range(len(list1)):
            ws[f"{list_row[i]}{j+9}"].value = list_group[i][j]
            ws["D6"].value = house
            ws["AD4"].value = dt_now.strftime('%Y/%m/%d %H:%M:%S')
            ws["AF22"].value = name
    rows = ws["A9": "AJ19"]
    for row in rows:
        values = [cell.coordinate for cell in row]
        for value in values: ws[value].font = font 
    output = BytesIO()
    wb.save(output)
    processed_data = output.getvalue()
    return processed_data